import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
import json
import random
import string
import openpyxl
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # 用于闪存消息
app.permanent_session_lifetime = timedelta(minutes=30)  # 设置 session 的过期时间

# 配置上传文件的目录和允许的扩展名
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 定义函数以检查文件扩展名是否允许
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload_product_image/<product_name>', methods=['GET', 'POST'])
def upload_product_image(product_name):
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('没有选择文件')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('没有选择文件')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            # 将文件名设置为产品名.jpg
            filename = secure_filename(f"{product_name}.jpg").replace(' ', '_')
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            # 更新Excel文件中的图片路径
            wb = openpyxl.load_workbook('inventory.xlsx')
            sheet = wb['Items']
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == product_name:  # B列是产品名称
                    row[3].value = filepath.replace('\\', '/')  # 假设D列是图片路径
                    break
            wb.save('inventory.xlsx')
            
            flash('封面图片上传成功')
            return redirect(url_for('product_record', product_name=product_name)) 
            
    return render_template('upload_product_image.html', product_name=product_name)

# 读取所有类别
def read_categories():
    wb = openpyxl.load_workbook('inventory.xlsx')
    sheet = wb['Items']
    categories = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        categories.add(row[0])  # A列是类别
    return list(categories)

def read_products_by_category(category):
    wb = openpyxl.load_workbook('inventory.xlsx')
    sheet = wb['Items']
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == category:  # 匹配类别
            product = {
                'name': row[1],  # 产品名称在 B 列
                'quantity': row[2],  # 数量在 C 列
                'image': row[3] if row[3] else 'static/uploads/default.jpg'  # D 列是照片路径
            }
            products.append(product)
    return products

# 读取某产品的进/出货记录
def read_product_records(product_name):
    wb = openpyxl.load_workbook('inventory.xlsx')
    sheet = wb['Shipment']
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == product_name:  # D列是产品名称
            record = {
                'date': row[0],  # A列是日期
                'recipient': row[1],  # B列是收货人
                'sender': row[2],  # C列是出货人
                'quantity': row[4],  # E列是数量
                'remaining': row[5] if len(row) > 5 else None  # F列是剩余库存（处理旧记录）
            }
            records.append(record)
    return records

# 首页展示类别
@app.route('/')
def home():
    categories = [
        {'name': 'CPU', 'image': 'static/uploads/cpu_image.jpg'},
        {'name': '主板', 'image': 'static/uploads/mb.jpg'},
        {'name': '内存', 'image': 'static/uploads/ram.jpg'},
        {'name': '显卡', 'image': 'static/uploads/gpu.jpg'},
        {'name': 'SSD', 'image': 'static/uploads/ssd.jpg'},
        {'name': '电源', 'image': 'static/uploads/psu.jpeg'},
        {'name': '机箱', 'image': 'static/uploads/case.jpg'},
        {'name': '风扇', 'image': 'static/uploads/fan.jpg'},
        {'name': '其他', 'image': 'static/uploads/others.jpg'},
        {'name': 'Laptop', 'image': 'static/uploads/laptop.jpg'},
        # 其他类别
    ]
    return render_template('home.html', categories=categories)


@app.route('/shipment_confirmation', methods=['POST'])
def shipment_confirmation():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "fail", "message": "Invalid JSON data"}), 415

        sender = data.get('sender')
        recipient = data.get('recipient')
        operation_time = data.get('operationTime')
        products = data.get('products', [])  # 使用 products 代替 items


        if not isinstance(products, list):  # 这里检查 products 是否为列表
            raise ValueError("Products must be a list")

        shipment_data = {
            'sender': sender,
            'recipient': recipient,
            'operation_time': operation_time,
            'products': products  # 更新字典键为 products
        }

        # 确保在渲染模板前打印出 shipment_data

        return render_template('shipment_confirmation.html', 
                               shipment_data=shipment_data,
                               senders=['Sender 1', 'Sender 2', 'Sender 3'], 
                               recipients=['Recipient 1', 'Recipient 2', 'Recipient 3'])

    except Exception as e:
        print(f"Error during rendering: {str(e)}")
        return jsonify({"status": "fail", "message": str(e)}), 500

@app.route('/shipment_overview')
def shipment_overview():
    categories = read_categories()
    products_by_category = {}
    
    for category in categories:
        products_by_category[category] = read_products_by_category(category)
    
    return render_template('shipment_overview.html', products_by_category=products_by_category)

@app.route('/process_shipment', methods=['POST'])
def process_shipment():
    try:
        data = request.get_json()
        sender = data.get('sender')
        recipient = data.get('recipient')
        operation_time = data.get('operationTime')
        products = data.get('products', [])

        # 打开并加载 Excel 文件
        wb = openpyxl.load_workbook('inventory.xlsx')
        items_sheet = wb['Items']
        shipment_sheet = wb['Shipment']

        # 遍历产品列表并更新库存
        for product in products:
            product_name = product['name']
            quantity_to_ship = product['quantity']

            for row in items_sheet.iter_rows(min_row=2):
                if row[1].value == product_name:
                    current_stock = row[2].value
                    if current_stock < quantity_to_ship:
                        return jsonify({"status": "fail", "message": f"库存不足: {product_name}"}), 400
                    
                    # 更新库存数量
                    row[2].value = current_stock - quantity_to_ship
                    break

            # 记录这次出货
            shipment_sheet.append([operation_time, recipient, sender, product_name, -quantity_to_ship, row[2].value])

        # 保存更新后的 Excel 文件
        wb.save('inventory.xlsx')

        return jsonify({"status": "success"}), 200

    except Exception as e:
        print(f"Error during shipment processing: {str(e)}")
        return jsonify({"status": "fail", "message": str(e)}), 500


# 展示某个类别的产品
@app.route('/category/<category_name>')
def show_category(category_name):
    products = read_products_by_category(category_name)
    return render_template('category.html', category_name=category_name, products=products)

# 展示某个产品的进/出货记录
@app.route('/product/<product_name>')
def product_record(product_name):
    records = read_product_records(product_name)
    return render_template('product_record.html', product_name=product_name, records=records)

@app.route('/add_stock', methods=['POST'])
def add_stock():
    product_name = request.form['product_name']
    quantity = int(request.form['quantity'])
    recipient = request.form['recipient']

    wb = openpyxl.load_workbook('inventory.xlsx')
    products_sheet = wb['Items']  # 修改为 products_sheet
    shipment_sheet = wb['Shipment']

    # 查找产品并更新库存数量
    for row in products_sheet.iter_rows(min_row=2):
        if row[1].value == product_name:
            current_stock = row[2].value
            if current_stock + quantity < 0:
                # 库存不能为负数，返回错误提示
                wb.close()
                return f"操作失败：库存不足。当前库存为 {current_stock} 件。", 400

            new_stock = current_stock + quantity
            row[2].value = new_stock
            break

    # 记录这次操作和剩余库存
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    operation_type = '进货' if quantity > 0 else '出货'
    shipment_sheet.append([current_time, recipient, "System", product_name, quantity, new_stock])

    # 保存更改到Excel文件
    wb.save('inventory.xlsx')

    return redirect(url_for('product_record', product_name=product_name))

if __name__ == '__main__':
    app.run(debug=True)
