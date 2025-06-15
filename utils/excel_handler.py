import openpyxl
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

class ExcelHandler:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.logger = logging.getLogger(__name__)
    
    def _load_workbook(self):
        """Load Excel workbook with error handling"""
        try:
            return openpyxl.load_workbook(self.excel_file)
        except FileNotFoundError:
            self.logger.error(f"Excel file {self.excel_file} not found")
            raise
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {str(e)}")
            raise
    
    def get_categories(self) -> List[str]:
        """Get all unique categories from inventory"""
        wb = self._load_workbook()
        sheet = wb['Items']
        categories = set()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Category in column A
                categories.add(row[0])
        
        wb.close()
        return sorted(list(categories))
    
    def get_products_by_category(self, category: Optional[str] = None) -> Dict[str, List[Dict]]:
        """Get products by category or all products grouped by category"""
        wb = self._load_workbook()
        sheet = wb['Items']
        
        if category:
            products = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == category:
                    product = {
                        'name': row[1] or 'Unknown',
                        'quantity': row[2] or 0,
                        'image': row[3] or 'static/uploads/default.jpg'
                    }
                    products.append(product)
            wb.close()
            return products
        else:
            products_by_category = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cat = row[0] or 'Uncategorized'
                if cat not in products_by_category:
                    products_by_category[cat] = []
                
                product = {
                    'name': row[1] or 'Unknown',
                    'quantity': row[2] or 0,
                    'image': row[3] or 'static/uploads/default.jpg'
                }
                products_by_category[cat].append(product)
            
            wb.close()
            return products_by_category
    
    def get_product_records(self, product_name: str) -> List[Dict]:
        """Get shipment records for a specific product"""
        wb = self._load_workbook()
        sheet = wb['Shipment']
        records = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == product_name:  # Product name in column D
                record = {
                    'date': row[0] or datetime.now(),
                    'recipient': row[1] or 'Unknown',
                    'sender': row[2] or 'System',
                    'quantity': row[4] or 0,
                    'remaining': row[5] if len(row) > 5 else None
                }
                records.append(record)
        
        wb.close()
        return sorted(records, key=lambda x: x['date'], reverse=True)
    
    def update_product_image(self, product_name: str, image_path: str) -> bool:
        """Update product image path in Excel"""
        try:
            wb = self._load_workbook()
            sheet = wb['Items']
            
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == product_name:  # Product name in column B
                    row[3].value = image_path.replace('\\', '/')  # Image path in column D
                    break
            
            wb.save(self.excel_file)
            wb.close()
            return True
        except Exception as e:
            self.logger.error(f"Error updating product image: {str(e)}")
            return False
    
    def update_stock(self, product_name: str, quantity: int, recipient: str) -> Dict[str, Any]:
        """Update product stock and record the transaction"""
        try:
            wb = self._load_workbook()
            items_sheet = wb['Items']
            shipment_sheet = wb['Shipment']
            
            # Find and update product stock
            product_found = False
            new_stock = 0
            
            for row in items_sheet.iter_rows(min_row=2):
                if row[1].value == product_name:  # Product name in column B
                    current_stock = row[2].value or 0
                    
                    if current_stock + quantity < 0:
                        wb.close()
                        return {
                            'success': False,
                            'message': f'库存不足。当前库存为 {current_stock} 件。'
                        }
                    
                    new_stock = current_stock + quantity
                    row[2].value = new_stock  # Update quantity in column C
                    product_found = True
                    break
            
            if not product_found:
                wb.close()
                return {
                    'success': False,
                    'message': f'产品 {product_name} 未找到。'
                }
            
            # Record the transaction
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            shipment_sheet.append([
                current_time,
                recipient,
                "System",
                product_name,
                quantity,
                new_stock
            ])
            
            wb.save(self.excel_file)
            wb.close()
            
            return {
                'success': True,
                'message': f'库存更新成功。新库存: {new_stock}'
            }
            
        except Exception as e:
            self.logger.error(f"Error updating stock: {str(e)}")
            return {
                'success': False,
                'message': f'更新失败: {str(e)}'
            }
    
    def process_shipment(self, sender: str, recipient: str, operation_time: str, products: List[Dict]) -> Dict[str, Any]:
        """Process a complete shipment with multiple products"""
        try:
            wb = self._load_workbook()
            items_sheet = wb['Items']
            shipment_sheet = wb['Shipment']
            
            # Validate stock availability first
            stock_issues = []
            for product in products:
                product_name = product['name']
                quantity_to_ship = product['quantity']
                
                for row in items_sheet.iter_rows(min_row=2):
                    if row[1].value == product_name:
                        current_stock = row[2].value or 0
                        if current_stock < quantity_to_ship:
                            stock_issues.append(f"{product_name} (需要: {quantity_to_ship}, 库存: {current_stock})")
                        break
            
            if stock_issues:
                wb.close()
                return {
                    'success': False,
                    'message': f'库存不足: {", ".join(stock_issues)}'
                }
            
            # Process the shipment
            for product in products:
                product_name = product['name']
                quantity_to_ship = product['quantity']
                
                # Update stock
                for row in items_sheet.iter_rows(min_row=2):
                    if row[1].value == product_name:
                        current_stock = row[2].value or 0
                        new_stock = current_stock - quantity_to_ship
                        row[2].value = new_stock
                        
                        # Record shipment
                        shipment_sheet.append([
                            operation_time,
                            recipient,
                            sender,
                            product_name,
                            -quantity_to_ship,
                            new_stock
                        ])
                        break
            
            wb.save(self.excel_file)
            wb.close()
            
            return {
                'success': True,
                'message': '出货成功'
            }
            
        except Exception as e:
            self.logger.error(f"Error processing shipment: {str(e)}")
            return {
                'success': False,
                'message': f'出货失败: {str(e)}'
            }
    
    def get_inventory_summary(self) -> Dict[str, Any]:
        """Get inventory summary statistics"""
        try:
            wb = self._load_workbook()
            sheet = wb['Items']
            
            total_products = 0
            total_value = 0
            categories = {}
            low_stock_items = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[1]:  # Product name exists
                    category = row[0] or 'Uncategorized'
                    quantity = row[2] or 0
                    
                    total_products += 1
                    
                    if category not in categories:
                        categories[category] = {'count': 0, 'total_quantity': 0}
                    
                    categories[category]['count'] += 1
                    categories[category]['total_quantity'] += quantity
                    
                    if quantity <= 5:  # Low stock threshold
                        low_stock_items.append({
                            'name': row[1],
                            'category': category,
                            'quantity': quantity
                        })
            
            wb.close()
            
            return {
                'total_products': total_products,
                'categories': categories,
                'low_stock_count': len(low_stock_items),
                'low_stock_items': low_stock_items
            }
            
        except Exception as e:
            self.logger.error(f"Error getting inventory summary: {str(e)}")
            return {'error': str(e)}
    
    def get_low_stock_items(self, threshold: int = 5) -> List[Dict]:
        """Get items with stock below threshold"""
        try:
            wb = self._load_workbook()
            sheet = wb['Items']
            low_stock_items = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[1] and (row[2] or 0) <= threshold:
                    low_stock_items.append({
                        'name': row[1],
                        'category': row[0] or 'Uncategorized',
                        'quantity': row[2] or 0,
                        'image': row[3] or 'static/uploads/default.jpg'
                    })
            
            wb.close()
            return low_stock_items
            
        except Exception as e:
            self.logger.error(f"Error getting low stock items: {str(e)}")
            return []